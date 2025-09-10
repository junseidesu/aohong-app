import os
import json
import customtkinter as ctk
import openpyxl as opx
from tkinter import filedialog, messagebox
import datetime

class Settings:
    # 永続化ファイル名のみ固定（その他は _DEFAULT_VALUES で一元管理）
    SETTINGS_FILE = "settings.json"
    
    # 以下の値（APP_NAME 含む UI/ファイル/行列番号 など）は _DEFAULT_VALUES のみで保持し
    # reset_to_defaults() によりクラス変数へ一括適用する。二重定義を避けるため
    # クラスレベルで個別代入しない。
    _DEFAULT_VALUES = {
        "PRICE_FILE_PATH": "フロンガス単価表2025.xlsx",
        "STOCK_FILE_PATH": "2025在庫.xlsx",
        "SALES_FILE_PATH": "R706 得意先別売上分析表.xlsx",
        "ID_ROW_IN_PRICE": 3,
        "PRICE_ROW_IN_PRICE": 25,
        "ID_COLUMN_IN_STOCK": 3,
        "PRICE_COLUMN_IN_STOCK": 10,
        "DATA_START_ROW_IN_STOCK": 7,
        "ID_COLUMN_IN_SALES": 4,
        "PROFIT_COLUMN_IN_SALES": 9,
        "PROFIT_RATE_COLUMN_IN_SALES": 10,
        "SALES_COLUMN_IN_SALES": 6,
        "SALES_NUM_COLUMN_IN_SALES": 8,
        "APP_NAME": "在庫単価転記アプリ",
        # 横幅を少し狭める（以前:400）
        "WINDOW_WIDTH": 340,
        # 高さもコンパクトに（以前:300）
        "WINDOW_HEIGHT": 240,
        "THEME": "light",
        "COLOR_THEME": "dark-blue",
        "ICON_FILE": "icon.ico"
    }
    
    @classmethod
    def load_settings(cls):
        """settings.json から設定を読み込みし、存在しない場合や欠損キーはデフォルト適用"""
        # まず全てデフォルトで初期化
        cls.reset_to_defaults()
        try:
            if not os.path.exists(cls.SETTINGS_FILE):
                return
            with open(cls.SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            # ファイルパス設定
            files = data.get("files", {})
            if files:
                cls.PRICE_FILE_PATH = files.get("price_file_path", cls.PRICE_FILE_PATH)
                cls.STOCK_FILE_PATH = files.get("stock_file_path", cls.STOCK_FILE_PATH)
                cls.SALES_FILE_PATH = files.get("sales_file_path", cls.SALES_FILE_PATH)
            # 行・列設定
            positions = data.get("positions", {})
            if positions:
                cls.ID_ROW_IN_PRICE = positions.get("id_row_in_price", cls.ID_ROW_IN_PRICE)
                cls.PRICE_ROW_IN_PRICE = positions.get("price_row_in_price", cls.PRICE_ROW_IN_PRICE)
                cls.ID_COLUMN_IN_STOCK = positions.get("id_column_in_stock", cls.ID_COLUMN_IN_STOCK)
                cls.PRICE_COLUMN_IN_STOCK = positions.get("price_column_in_stock", cls.PRICE_COLUMN_IN_STOCK)
                cls.DATA_START_ROW_IN_STOCK = positions.get("data_start_row_in_stock", cls.DATA_START_ROW_IN_STOCK)
                cls.ID_COLUMN_IN_SALES = positions.get("id_column_in_sales", cls.ID_COLUMN_IN_SALES)
                cls.PROFIT_COLUMN_IN_SALES = positions.get("profit_column_in_sales", cls.PROFIT_COLUMN_IN_SALES)
                cls.PROFIT_RATE_COLUMN_IN_SALES = positions.get("profit_rate_column_in_sales", cls.PROFIT_RATE_COLUMN_IN_SALES)
                cls.SALES_COLUMN_IN_SALES = positions.get("sales_column_in_sales", cls.SALES_COLUMN_IN_SALES)
                cls.SALES_NUM_COLUMN_IN_SALES = positions.get("sales_num_column_in_sales", cls.SALES_NUM_COLUMN_IN_SALES)
            print("設定を読み込みました")
        except Exception as e:
            print(f"設定の読み込みエラー: {e}")
    
    @classmethod
    def save_settings(cls):
        """現在の設定をsettings.jsonに保存"""
        try:
            settings_data = {
                "files": {
                    "price_file_path": cls.PRICE_FILE_PATH,
                    "stock_file_path": cls.STOCK_FILE_PATH,
                    "sales_file_path": cls.SALES_FILE_PATH
                },
                "positions": {
                    "id_row_in_price": cls.ID_ROW_IN_PRICE,
                    "price_row_in_price": cls.PRICE_ROW_IN_PRICE,
                    "id_column_in_stock": cls.ID_COLUMN_IN_STOCK,
                    "price_column_in_stock": cls.PRICE_COLUMN_IN_STOCK,
                    "data_start_row_in_stock": cls.DATA_START_ROW_IN_STOCK,
                    "id_column_in_sales": cls.ID_COLUMN_IN_SALES,
                    "profit_column_in_sales": cls.PROFIT_COLUMN_IN_SALES,
                    "profit_rate_column_in_sales": cls.PROFIT_RATE_COLUMN_IN_SALES,
                    "sales_column_in_sales": cls.SALES_COLUMN_IN_SALES,
                    "sales_num_column_in_sales": cls.SALES_NUM_COLUMN_IN_SALES
                }
            }
            
            with open(cls.SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(settings_data, f, ensure_ascii=False, indent=2)
            
            print("設定を保存しました")
            return True
        except Exception as e:
            print(f"設定の保存エラー: {e}")
            return False
        
    @classmethod
    def reset_to_defaults(cls):
        """設定をデフォルト値にリセット（_DEFAULT_VALUES を唯一のソースとする）"""
        try:
            for key, value in cls._DEFAULT_VALUES.items():
                setattr(cls, key, value)
            return True
        except Exception as e:
            print(f"設定リセットエラー: {e}")
            return False
    
    @classmethod
    def get_default_value(cls, key):
        """指定されたキーのデフォルト値を取得"""
        return cls._DEFAULT_VALUES.get(key, None)
    
    @classmethod
    def is_default_value(cls, key):
        """現在の値がデフォルト値かどうかを確認"""
        current_value = getattr(cls, key, None)
        default_value = cls.get_default_value(key)
        return current_value == default_value

# 初期化: デフォルト → 上書き読み込み
Settings.load_settings()

class SettingsWindow(ctk.CTkToplevel):
    def __init__(self, app):
        super().__init__()
        self.app = app  # 親アプリ参照
        self.title("設定")
        # 初期最小サイズ (内容に応じて後で拡張)
        self.minsize(420, 260)
        self.resizable(True, True)
        self._dirty = False
        self._flash_after_ids = []

        try:
            self.iconbitmap(Settings.ICON_FILE)
        except Exception:
            pass
        # スクロール可能領域（全体コンテンツ）
        self.scroll = ctk.CTkScrollableFrame(self, label_text="設定", fg_color=("#E0E0E0", "#1a1a1a"))
        self.scroll.pack(fill="both", expand=True, padx=10, pady=10)

        # ファイル設定セクション
        file_section = ctk.CTkFrame(self.scroll, fg_color=("#D0D0D0", "#0f0f0f"))
        file_section.pack(fill="x", pady=(0,10))
        ctk.CTkLabel(file_section, text="ファイル設定", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=6, sticky="w", pady=(4,4))

        # 共通: パス Entry + ボタン (参照/開く/クリア)
        self.file_vars = {
            "price": ctk.StringVar(value=Settings.PRICE_FILE_PATH),
            "stock": ctk.StringVar(value=Settings.STOCK_FILE_PATH),
            "sales": ctk.StringVar(value=Settings.SALES_FILE_PATH),
        }
        file_labels = {"price": "単価表", "stock": "在庫表", "sales": "売上表"}
        for r, key in enumerate(["price", "stock", "sales"], start=1):
            ctk.CTkLabel(file_section, text=file_labels[key], width=70, anchor="w").grid(row=r, column=0, padx=4, pady=2, sticky="w")
            entry = ctk.CTkEntry(file_section, textvariable=self.file_vars[key], width=230)
            entry.grid(row=r, column=1, padx=4, pady=2, sticky="we", columnspan=2)
            browse_btn = ctk.CTkButton(file_section, text="参照", width=50, command=lambda k=key: self.select_file(k))
            browse_btn.grid(row=r, column=3, padx=2, pady=2)
            open_btn = ctk.CTkButton(file_section, text="開く", width=50, command=lambda k=key: self.open_in_explorer(k))
            open_btn.grid(row=r, column=4, padx=2, pady=2)
            clear_btn = ctk.CTkButton(file_section, text="×", width=30, command=lambda k=key: self.clear_path(k))
            clear_btn.grid(row=r, column=5, padx=2, pady=2)
        for c in range(0,6):
            file_section.grid_columnconfigure(c, weight= (1 if c in (1,2) else 0))

        # 行・列設定編集フレーム
        self.position_frame = ctk.CTkFrame(self.scroll, fg_color=("#D0D0D0", "#0f0f0f"))
        self.position_frame.pack(pady=5, padx=0, fill="x")
        ctk.CTkLabel(self.position_frame, text="行・列設定", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=4, pady=(4,4), sticky="w")

        # 編集対象キーと日本語ラベル
        self.position_keys = [
            ("ID_ROW_IN_PRICE", "単価表: ID 行"),
            ("PRICE_ROW_IN_PRICE", "単価表: 単価 行"),
            ("ID_COLUMN_IN_STOCK", "在庫表: ID 列"),
            ("PRICE_COLUMN_IN_STOCK", "在庫表: 単価 列"),
            ("DATA_START_ROW_IN_STOCK", "在庫表: データ開始 行"),
            ("ID_COLUMN_IN_SALES", "売上表: ID 列"),
            ("PROFIT_COLUMN_IN_SALES", "売上表: 利益 列"),
            ("PROFIT_RATE_COLUMN_IN_SALES", "売上表: 利益率 列"),
            ("SALES_COLUMN_IN_SALES", "売上表: 売上金額 列"),
            ("SALES_NUM_COLUMN_IN_SALES", "売上表: 売上数量 列"),
        ]
        # 数値入力 validate
        def _only_int(P):
            return P.isdigit() or P == ""
        vcmd = (self.register(_only_int), "%P")
        self.position_entries = {}
        for idx, (key, label_text) in enumerate(self.position_keys, start=1):
            ctk.CTkLabel(self.position_frame, text=label_text, anchor="w").grid(row=idx, column=0, padx=4, pady=1, sticky="w")
            var = ctk.StringVar(value=str(getattr(Settings, key)))
            entry = ctk.CTkEntry(self.position_frame, textvariable=var, width=70, validate="key", validatecommand=vcmd)
            entry.grid(row=idx, column=1, padx=4, pady=1, sticky="w")
            self.position_entries[key] = (entry, var)
        self.position_frame.grid_columnconfigure(0, weight=1)

        # 現在の設定表示
        self.info_frame = ctk.CTkFrame(master=self.scroll, fg_color=("#D0D0D0", "#0f0f0f"))
        self.info_frame.pack(pady=8, padx=0, fill="x")

        ctk.CTkLabel(self.info_frame, text="現在の設定", font=("Arial", 14, "bold")).pack(pady=5)

        self.price_label = ctk.CTkLabel(self.info_frame, text=f"単価表: {os.path.basename(Settings.PRICE_FILE_PATH)}")
        self.price_label.pack(pady=1)
        self.stock_label = ctk.CTkLabel(self.info_frame, text=f"在庫表: {os.path.basename(Settings.STOCK_FILE_PATH)}")
        self.stock_label.pack(pady=1)
        self.sales_label = ctk.CTkLabel(self.info_frame, text=f"売上表: {os.path.basename(Settings.SALES_FILE_PATH)}")
        self.sales_label.pack(pady=1)

        # 保存ボタン
        self.save_button = ctk.CTkButton(
            master=self.scroll,
            text="設定を保存",
            command=self.save_settings
        )
        self.save_button.pack(pady=10)

        self.reset_button = ctk.CTkButton(
            master=self.scroll,
            text="設定をリセット",
            command=self.reset_settings
        )
        self.reset_button.pack(pady=10)

        # --- 最前面＆フォーカス処理 ---
        self.after(0, self._bring_to_front)
        # 起動後に内容サイズへフィット
        self.after(80, self.auto_fit_size)

        # 変更検知
        for key, (entry, var) in self.position_entries.items():
            var.trace_add("write", lambda *_args, k=key: self.mark_dirty(k))
        for k, var in self.file_vars.items():
            var.trace_add("write", lambda *_a, kk=k: self.mark_dirty(kk))
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def _bring_to_front(self):
        try:
            self.attributes("-topmost", True)
            self.lift()
            self.focus_force()
            # 数百 ms 後に topmost を解除
            self.after(300, lambda: self.attributes("-topmost", False))
        except Exception:
            pass

    def select_file(self, file_type):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")]
        )
        if file_path:
            self.file_vars[file_type].set(file_path)
            self.update_file_labels()
            self.auto_fit_size(only_expand=False)

    def open_in_explorer(self, key):
        path = self.file_vars[key].get().strip()
        if not path:
            return
        try:
            if os.path.exists(path):
                os.startfile(os.path.dirname(path) if os.path.isfile(path) else path)
        except Exception as e:
            messagebox.showerror("エラー", f"エクスプローラを開けません: {e}")

    def clear_path(self, key):
        self.file_vars[key].set("")
        self.update_file_labels()

    def save_settings(self):
        # まず行・列の入力値を検証し Settings へ反映
        try:
            for key, (_entry, var) in self.position_entries.items():
                raw = var.get().strip()
                if raw == "":
                    raise ValueError(f"{key} が空です")
                value = int(raw)
                if value <= 0:
                    raise ValueError(f"{key} は正の整数である必要があります")
                setattr(Settings, key, value)
        except ValueError as ve:
            messagebox.showerror("入力エラー", str(ve))
            return

        # ファイルパス適用
        Settings.PRICE_FILE_PATH = self.file_vars["price"].get().strip() or Settings.PRICE_FILE_PATH
        Settings.STOCK_FILE_PATH = self.file_vars["stock"].get().strip() or Settings.STOCK_FILE_PATH
        Settings.SALES_FILE_PATH = self.file_vars["sales"].get().strip() or Settings.SALES_FILE_PATH
        self.update_file_labels()

        if Settings.save_settings():
            messagebox.showinfo("設定保存", "設定を保存しました")
            if self.app:
                try:
                    self.app.refresh_settings_ui()
                except Exception:
                    pass
            # UI サイズ再調整
            self.auto_fit_size(only_expand=True)
            self._dirty = False
            self.flash_saved()
    
    def reset_settings(self):
        """設定をデフォルトへ戻し UI を更新"""
        Settings.reset_to_defaults()
        Settings.save_settings()
        messagebox.showinfo("設定リセット", "設定をリセットしました")
        # 数値エントリ更新
        for key, (_, var) in self.position_entries.items():
            var.set(str(getattr(Settings, key)))
        # ファイルパス表示更新
        self.file_vars["price"].set(Settings.PRICE_FILE_PATH)
        self.file_vars["stock"].set(Settings.STOCK_FILE_PATH)
        self.file_vars["sales"].set(Settings.SALES_FILE_PATH)
        self.update_file_labels()
        # サイズ調整
        self.auto_fit_size(only_expand=True)
        self._dirty = False

    def auto_fit_size(self, extra_w: int = 4, extra_h: int = 10, only_expand: bool = False):
        """現在内容の要求サイズに合わせてウィンドウを調整。
        only_expand=True の場合は縮小せず拡張だけ行う。"""
        try:
            self.update_idletasks()
            req_w = self.winfo_reqwidth() + extra_w
            req_h = self.winfo_reqheight() + extra_h
            # 上限設定（画面 80%）
            sw = self.winfo_screenwidth()
            sh = self.winfo_screenheight()
            max_w = int(sw * 0.8)
            max_h = int(sh * 0.8)
            cur_w = self.winfo_width()
            cur_h = self.winfo_height()
            if cur_w <= 1 or cur_h <= 1:  # 初期取得失敗時フォールバック
                cur_w, cur_h = 500, 400
            if only_expand:
                new_w = max(req_w, cur_w)
                new_h = max(req_h, cur_h)
            else:
                new_w, new_h = req_w, req_h
            new_w = min(new_w, max_w)
            new_h = min(new_h, max_h)
            x = self.winfo_x()
            y = self.winfo_y()
            self.geometry(f"{new_w}x{new_h}+{x}+{y}")
        except Exception as e:
            print(f"SettingsWindow auto_fit_size エラー: {e}")

    # 付加機能: 状態/ハイライト
    def update_file_labels(self):
        self.price_label.configure(text=f"単価表: {os.path.basename(self.file_vars['price'].get()) or '-'}")
        self.stock_label.configure(text=f"在庫表: {os.path.basename(self.file_vars['stock'].get()) or '-'}")
        self.sales_label.configure(text=f"売上表: {os.path.basename(self.file_vars['sales'].get()) or '-'}")

    def mark_dirty(self, key):
        self._dirty = True
        # 変更された数値エントリ背景軽く強調
        if key in dict(self.position_keys):
            entry, var = self.position_entries[key]
            default = str(Settings.get_default_value(key))
            if var.get() and var.get() != default:
                entry.configure(fg_color=("#444444", "#444444"))
            else:
                entry.configure(fg_color=("#333333", "#333333"))

    def flash_saved(self):
        for entry, _ in self.position_entries.values():
            orig = entry.cget("fg_color")
            entry.configure(fg_color=("#155b2f", "#155b2f"))
            self.after(350, lambda e=entry, o=orig: e.configure(fg_color=o))

    def on_close(self):
        if self._dirty:
            if not messagebox.askyesno("確認", "未保存の変更があります。閉じてもよいですか？"):
                return
        self.destroy()

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        # リサイズ許可
        self.resizable(True, True)

        # テーマ & 基本属性
        ctk.set_appearance_mode(Settings.THEME)
        ctk.set_default_color_theme(Settings.COLOR_THEME)
        self.title(Settings.APP_NAME)
        self.minsize(Settings.WINDOW_WIDTH, Settings.WINDOW_HEIGHT)

        try:
            self.iconbitmap(Settings.ICON_FILE)
        except Exception:
            pass

        # 上部: 設定ボタン
        self.file_frame = ctk.CTkFrame(self, fg_color=("#E0E0E0", "#1a1a1a"))
        self.file_frame.pack(padx=8, pady=2, fill="x")
        self.settings_button = ctk.CTkButton(self.file_frame, text="設定", width=70, command=self.open_settings)
        self.settings_button.pack(pady=2)

        # メインフレーム
        self.frame = ctk.CTkFrame(self, fg_color=("#E0E0E0", "#1a1a1a"))
        self.frame.pack(pady=6, padx=10, fill="both", expand=True)

        # 年月選択
        now = datetime.datetime.now()
        self.year_var = ctk.StringVar(value=f"{now.year}年")
        self.month_var = ctk.StringVar(value=f"{now.month}月")
        self.year_month_menu_frame = ctk.CTkFrame(self.frame, fg_color=("#D0D0D0", "#0f0f0f"))
        self.year_month_menu_frame.pack(pady=2, fill="x")
        current_year = now.year
        year_options = [f"{i}年" for i in range(current_year-1, current_year+2)]
        self.year_menu = ctk.CTkOptionMenu(self.year_month_menu_frame, variable=self.year_var, values=year_options, width=90)
        self.year_menu.grid(row=0, column=0, padx=(4,4), pady=2)
        self.year_menu.set(f"{current_year}年")
        month_options = [f"{i}月" for i in range(1, 13)]
        self.month_menu = ctk.CTkOptionMenu(self.year_month_menu_frame, variable=self.month_var, values=month_options, width=68)
        self.month_menu.grid(row=0, column=1, padx=(4,4), pady=2)
        self.month_menu.set(f"{now.month}月")
        self.year_month_menu_frame.grid_columnconfigure(0, weight=1)
        self.year_month_menu_frame.grid_columnconfigure(1, weight=1)

        # 操作ボタン
        self.button_1 = ctk.CTkButton(self.frame, text="在庫単価を在庫表に転記", command=self.update_stock_list, width=200)
        self.button_1.pack(pady=(16,8))
        self.button_2 = ctk.CTkButton(self.frame, text="在庫単価を売上表に転記", command=self.update_sales_list, width=200)
        self.button_2.pack(pady=(8,16))

        # 進行状況バー & ステータス
        self.progress = ctk.CTkProgressBar(self.frame)
        self.progress.pack(pady=(1, 3), fill="x")
        self.progress.set(0)
        self.status_var = ctk.StringVar(value="準備完了")
        self.status_label = ctk.CTkLabel(self, textvariable=self.status_var, anchor="w")
        self.status_label.pack(fill="x", side="bottom")

        # 初期フィット（幅・高さをできるだけ詰める）
        self.after(100, lambda: self.auto_fit_size(extra_w=2, extra_h=2, only_expand=False))

    def open_settings(self):
        """設定ウィンドウを開く"""
        # 既に開いている場合は再利用して前面へ
        if hasattr(self, "settings_window") and self.settings_window.winfo_exists():
            try:
                self.settings_window._bring_to_front()
            except Exception:
                pass
            return
        # 新規作成
        self.settings_window = SettingsWindow(self)
        try:
            # 親との関連付け（Windows でタスクバー分離を防ぐ）
            self.settings_window.transient(self)
            # モーダル風（他操作をブロックしたい場合） ※不要ならコメントアウト
            # self.settings_window.grab_set()
        except Exception:
            pass

    def refresh_settings_ui(self):
        """Settings の変更内容をメインウィンドウへ反映"""
        try:
            # タイトル・サイズ
            self.title(Settings.APP_NAME)
            # 最小サイズを更新（ユーザーの手動リサイズを尊重するため geometry 直接設定は避ける）
            self.minsize(Settings.WINDOW_WIDTH, Settings.WINDOW_HEIGHT)
            # テーマ適用（appearance / color）
            ctk.set_appearance_mode(Settings.THEME)
            ctk.set_default_color_theme(Settings.COLOR_THEME)
            # アイコン
            try:
                self.iconbitmap(Settings.ICON_FILE)
            except Exception:
                pass
            # コンテンツに合わせ再フィット（現在サイズが小さすぎる場合のみ拡張）
            self.auto_fit_size(only_expand=True)
        except Exception as e:
            print(f"UI反映エラー: {e}")

    def auto_fit_size(self, extra_w: int = 6, extra_h: int = 4, only_expand: bool = False):
        """現在の要求サイズに合わせてウィンドウを調整。

        extra_w / extra_h: 余白を加算
        only_expand: True の場合、現在サイズより小さくはしない（ユーザーの手動拡大を維持）
        """
        try:
            self.update_idletasks()
            req_w = self.winfo_reqwidth() + extra_w
            req_h = self.winfo_reqheight() + extra_h
            sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
            max_w = int(sw * 0.85); max_h = int(sh * 0.85)
            cur_w = self.winfo_width()
            cur_h = self.winfo_height()
            # 起動直後は width/height が 1 のことがあるのでフォールバック
            if cur_w <= 1 or cur_h <= 1:
                cur_w, cur_h = Settings.WINDOW_WIDTH, Settings.WINDOW_HEIGHT
            if only_expand:
                new_w = max(req_w, cur_w)
                new_h = max(req_h, cur_h)
            else:
                new_w, new_h = req_w, req_h
            new_w = min(new_w, max_w); new_h = min(new_h, max_h)
            # 位置は維持
            x = self.winfo_x()
            y = self.winfo_y()
            self.geometry(f"{new_w}x{new_h}+{x}+{y}")
        except Exception as e:
            print(f"auto_fit_size エラー: {e}")

    def update_stock_list(self):
        # ボタン無効化 & 進捗初期化
        self._start_long_task("在庫単価更新中...")
        # Settings から直接ファイルパスを取得
        try:
            price_list = opx.load_workbook(Settings.PRICE_FILE_PATH, data_only=True)
            stock_list = opx.load_workbook(Settings.STOCK_FILE_PATH, data_only=False)
        except FileNotFoundError as e:
            messagebox.showerror("エラー", f"ファイルが見つかりません: {e}")
            return self._end_long_task()
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました: {e}")
            return self._end_long_task()

        try:
            price_sheet = price_list["一般総平均"]
        except KeyError:
            messagebox.showerror("エラー", "一般総平均シートが見つかりません")
            return self._end_long_task()
        
        selected_year, selected_month = self.get_selected_year_month()
        selected_year_int = int(selected_year[:-1])
        selected_month_int = int(selected_month[:-1])
        converted_year_month = f"{selected_year_int}{selected_month_int:02}"
        
        month_cell = self.search_string_in_row(price_sheet, 1, converted_year_month)
        next_month_cell = self.search_string_in_row(price_sheet, 1, self.calculate_calendar(selected_year_int, selected_month_int, 1))
        
        if not month_cell:
            messagebox.showerror("エラー", f"{converted_year_month}のセルが見つかりません")
            return self._end_long_task()
        
        id_price_dict = {}
        next_month_col = None
        if next_month_cell:
            next_month_col = next_month_cell.column
        else:
            for i in range(1, price_sheet.max_column+2):
                num_of_none = 0
                for j in range(1, 26):
                    if price_sheet.cell(row=j, column=i).value is None:
                        num_of_none += 1
                if num_of_none == 25:
                    next_month_col = i
                    break
        
        # Settings から行・列番号を直接取得
        for i in range(month_cell.column, next_month_col):
            id_cell = price_sheet.cell(row=Settings.ID_ROW_IN_PRICE, column=i)
            price_cell = price_sheet.cell(row=Settings.PRICE_ROW_IN_PRICE, column=i)
            if id_cell.value and price_cell.value and price_cell.value!=0:
                id_price_dict[id_cell.value] = price_cell.value
                
        # 在庫表に転記
        sheetnames = stock_list.sheetnames
        stock_sheet = None
        for sheetname in sheetnames:
            if converted_year_month in sheetname:
                stock_sheet = stock_list[sheetname]
                break
        
        if not stock_sheet:
            messagebox.showerror("エラー", f"{converted_year_month}の在庫シートが見つかりません")
            return self._end_long_task()

        fix_point_count = 0
        for row_num in range(1, stock_sheet.max_row+1):
            id = stock_sheet.cell(row=row_num, column=Settings.ID_COLUMN_IN_STOCK).value
            price = id_price_dict.get(id, None)
            if price:
                stock_sheet.cell(row=row_num, column=Settings.PRICE_COLUMN_IN_STOCK, value=price)
                fix_point_count += 1
                print(f"ID: {id}, Price: {price} updated")
            if row_num % 50 == 0 or row_num == stock_sheet.max_row:
                try:
                    self.progress.set(row_num/stock_sheet.max_row)
                    self.status_var.set(f"在庫処理 {row_num}/{stock_sheet.max_row}")
                    self.update_idletasks()
                except Exception:
                    pass

        try:
            stock_list.save(Settings.STOCK_FILE_PATH)
            messagebox.showinfo("成功", f"{fix_point_count}件の価格を更新しました")
            self.status_var.set(f"在庫更新完了 {fix_point_count}件")
        except PermissionError:
            messagebox.showerror("エラー", "在庫ファイルが開かれています。閉じてから再度実行してください。")
        except Exception as e:
            messagebox.showerror("エラー", f"予期しないエラーが発生しました: {e}")
        finally:
            self._end_long_task()

    def update_sales_list(self):
        self._start_long_task("売上更新中...")
        try:
            stock_list = opx.load_workbook(Settings.STOCK_FILE_PATH, data_only=True)
            sales_list = opx.load_workbook(Settings.SALES_FILE_PATH, data_only=True)
        except FileNotFoundError as e:
            messagebox.showerror("エラー", f"ファイルが見つかりません: {e}")
            return self._end_long_task()
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました: {e}")
            return self._end_long_task()

        selected_year, selected_month = self.get_selected_year_month()
        selected_year_int = int(selected_year[:-1])
        selected_month_int = int(selected_month[:-1])
        converted_year_month = f"{selected_year_int}{selected_month_int:02}"
        
        sheetnames = stock_list.sheetnames
        stock_sheet = None
        for sheetname in sheetnames:
            if converted_year_month in sheetname:
                stock_sheet = stock_list[sheetname]
                break
        
        if not stock_sheet:
            messagebox.showerror("エラー", f"{converted_year_month}の在庫シートが見つかりません")
            return self._end_long_task()

        sales_sheet = sales_list.active

        id_price_dict = {}
        for row in range(Settings.DATA_START_ROW_IN_STOCK, stock_sheet.max_row + 1):
            id = stock_sheet.cell(row=row, column=Settings.ID_COLUMN_IN_STOCK).value
            price = stock_sheet.cell(row=row, column=Settings.PRICE_COLUMN_IN_STOCK).value
            if id and price:
                id_price_dict[id] = price

        fix_point_count = 0
        for row in range(1, sales_sheet.max_row+1):
            id = sales_sheet.cell(row=row, column=Settings.ID_COLUMN_IN_SALES).value
            price = id_price_dict.get(id, None)
            if price:
                try:
                    sales_value = sales_sheet.cell(row=row, column=Settings.SALES_COLUMN_IN_SALES).value
                    sales_num_value = sales_sheet.cell(row=row, column=Settings.SALES_NUM_COLUMN_IN_SALES).value

                    if sales_value is not None and sales_num_value is not None:
                        sales = float(sales_value)
                        sales_num = float(sales_num_value)

                        if sales and sales_num and sales != 0:
                            profit = sales - sales_num * price
                            profit_rate = profit / sales  # 利益率
                            sales_sheet.cell(row=row, column=Settings.PROFIT_COLUMN_IN_SALES, value=profit)
                            sales_sheet.cell(row=row, column=Settings.PROFIT_RATE_COLUMN_IN_SALES, value=profit_rate)
                            fix_point_count += 1
                    if row % 100 == 0 or row == sales_sheet.max_row:
                        try:
                            self.progress.set(row/sales_sheet.max_row)
                            self.status_var.set(f"売上処理 {row}/{sales_sheet.max_row}")
                            self.update_idletasks()
                        except Exception:
                            pass
                    print(f"ID: {id}, Price: {price} updated")
                except (ValueError, TypeError, ZeroDivisionError) as e:
                    print(f"行 {row} でデータ変換エラー: {e}")
                    continue

        try:
            sales_list.save(Settings.SALES_FILE_PATH)
            messagebox.showinfo("成功", f"{fix_point_count}件の売上データを更新しました")
            self.status_var.set(f"売上更新完了 {fix_point_count}件")
        except PermissionError:
            messagebox.showerror("エラー", "売上ファイルが開かれています。閉じてから再度実行してください。")
        except Exception as e:
            messagebox.showerror("エラー", f"予期しないエラーが発生しました: {e}")
        finally:
            self._end_long_task()

    # 長時間処理補助
    def _start_long_task(self, status_msg: str):
        try:
            self.button_1.configure(state="disabled")
            self.button_2.configure(state="disabled")
            self.progress.set(0)
            self.status_var.set(status_msg)
            self.configure(cursor="watch")
            self.update_idletasks()
        except Exception:
            pass

    def _end_long_task(self):
        try:
            self.button_1.configure(state="normal")
            self.button_2.configure(state="normal")
            self.configure(cursor="")
            self.progress.set(0)
        except Exception:
            pass

    def search_string_in_row(self, sheet, row, search_string):
        """指定されたExcelシートの特定の行から、指定された文字列を含むセルを検索する関数"""
        target_cell = None
        for row_data in sheet.iter_rows(min_row=row, max_row=row):
            for cell in row_data:
                if cell.value is not None:
                    try:
                        if search_string in str(cell.value):
                            target_cell = cell
                            break
                    except Exception:
                        continue
            if target_cell:
                break
        return target_cell
    
    def get_selected_year_month(self):
        selected_year = self.year_var.get()
        selected_month = self.month_var.get()
        return selected_year, selected_month

    def calculate_calendar(self, year, month, gap):
        if month + gap < 13 and month + gap > 0:
            return f"{year}{month+gap:02}"
        elif month + gap > 12:
            return f"{year+1}{month+gap-12:02}"
        else:
            return f"{year-1}{month+gap+12:02}"

if __name__ == "__main__":
    app = App()
    app.mainloop()